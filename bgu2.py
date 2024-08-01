import openpyxl, asyncio
from main import DBCommands

db = DBCommands()

path = ('/home/BLOOD.LOCAL/admin_sz/Projects/test_uploads/')

exel = openpyxl.load_workbook(path+'Сведения о пользователях БГУ2.0.xlsx')
bgu2 = exel['Лист_1']


async def verify_bgu2_permissions():
    for number_row in range(1, len(bgu2['A'])):
            worker_id = await db.get_worker_id(str(bgu2.cell(row=number_row, column=2).value))
            if bgu2.cell(row=number_row, column=4).value == 'Да' and len(worker_id)>0:
                await db.plus_bgu2(int(worker_id[0][0]))
            elif bgu2.cell(row=number_row, column=4).value != 'Да' and len(worker_id)>0:                
                await db.del_bgu2(int(worker_id[0][0]))


if __name__ == "__main__":
    asyncio.run(verify_bgu2_permissions())