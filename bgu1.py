import openpyxl, asyncio
from main import DBCommands

db = DBCommands()

path = ('/home/BLOOD.LOCAL/admin_sz/Projects/test_uploads/')

exel = openpyxl.load_workbook(path+'Пользователи БГУ1.0.xlsx')
bgu1 = exel['Лист1']


async def verify_bgu1_permissions():
    for number_row in range(1, len(bgu1['A'])):
            worker_id = await db.get_worker_id(str(bgu1.cell(row=number_row, column=2).value))
            if bgu1.cell(row=number_row, column=3).value == 'Да ' and len(worker_id)>0:
                await db.plus_bgu1(int(worker_id[0][0]))
            elif bgu1.cell(row=number_row, column=3).value != 'Да ' and len(worker_id)>0:                
                await db.del_bgu1(int(worker_id[0][0]))


if __name__ == "__main__":
    asyncio.run(verify_bgu1_permissions())