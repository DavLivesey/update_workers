import openpyxl, asyncio
from main import DBCommands

db = DBCommands()

path = ('/home/BLOOD.LOCAL/admin_sz/Projects/test_uploads/')

exel = openpyxl.load_workbook(path+'SED.xlsx')
sed = exel['Export']


async def verify_sed_permissions():
    for number_row in range(1, len(sed['A'])):
            worker_id = await db.get_worker_id_with_ad(sed.cell(row=number_row, column=1).value)
            if sed.cell(row=number_row, column=6).value == 'Да' and len(worker_id)>0:
                await db.plus_SED(int(worker_id[0][0]))
            elif sed.cell(row=number_row, column=6).value != 'Да' and len(worker_id)>0:            
                await db.del_SED(int(worker_id[0][0]))


if __name__ == "__main__":
    asyncio.run(verify_sed_permissions())