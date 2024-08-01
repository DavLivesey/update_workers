import openpyxl, asyncio
from main import DBCommands

db = DBCommands()

path = ('/home/BLOOD.LOCAL/admin_sz/Projects/test_uploads/')

exel = openpyxl.load_workbook(path+'Сведения о пользователях ЗКГУ.xlsx')
zkgu = exel['Лист_1']


async def verify_zkgu_permissions():
    for number_row in range(1, len(zkgu['A'])):
            worker_id = await db.get_worker_id(zkgu.cell(row=number_row, column=2).value)
            if zkgu.cell(row=number_row, column=7).value == 'Да' and len(worker_id)>0:                
                await db.plus_zkgu(int(worker_id[0][0]))
            elif zkgu.cell(row=number_row, column=7).value != 'Да' and len(worker_id)>0:                    
                await db.del_zkgu(int(worker_id[0][0]))


if __name__ == "__main__":
    asyncio.run(verify_zkgu_permissions())