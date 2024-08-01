import openpyxl, asyncio
from main import DBCommands

db = DBCommands()

path = ('/home/wsadmin/uploads/bgu1')

exel = openpyxl.load_workbook(path+'Пользователи Диетпитание.xlsx')
dieta = exel['Лист_1']


async def verify_dieta_permissions():
    for number_row in range(1, len(dieta['A'])):
            worker_id = await db.get_worker_id(str(dieta.cell(row=number_row, column=2).value))
            if dieta.cell(row=number_row, column=4).value == 'Да' and len(worker_id)>0:
                await db.plus_dieta(int(worker_id[0][0]))
            elif dieta.cell(row=number_row, column=4).value != 'Да' and len(worker_id)>0:                
                await db.del_dieta(int(worker_id[0][0]))


if __name__ == "__main__":
    asyncio.run(verify_dieta_permissions())
