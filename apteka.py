import openpyxl, asyncio
from main import DBCommands

db = DBCommands()

path = ('/home/BLOOD.LOCAL/admin_sz/Projects/test_uploads/')

exel = openpyxl.load_workbook(path+'Сведения о пользователях Аптека.xlsx')
apteka = exel['TDSheet']

async def verify_apteka_permissions():
    for number_row in range(1, len(apteka['A'])):
        worker_id = await db.get_worker_id(apteka.cell(row=number_row, column=2).value)
        if apteka.cell(row=number_row, column=7).value == 'Да' and len(worker_id)>0:                
            await db.plus_apteka(int(worker_id[0][0]))
        elif apteka.cell(row=number_row, column=7).value != 'Да' and len(worker_id)>0:            
            await db.del_apteka(int(worker_id[0][0]))


if __name__ == "__main__":
    asyncio.run(verify_apteka_permissions())
